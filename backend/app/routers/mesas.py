import csv
import io
from datetime import date, datetime, time, timedelta
from types import SimpleNamespace

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import Row, case, func, select
from sqlalchemy.exc import IntegrityError, NoResultFound
from sqlalchemy.ext.asyncio import AsyncSession

from app.database import get_session
from app.models import CategoriaMesa, Mesa, ResolutorMesa, SolicitanteMesa, VentanaMesa
from app.schemas import (
    ExtraccionMesa,
    GrupoSoporte,
    MesaCerrar,
    MesaCreate,
    MesaOut,
    MesaUpdate,
    PaginaMesas,
    PanelMesasKPIs,
    ReporteMesasSemanal,
)
from app.services.clustering import agrupar_por_similitud, tema_representativo
from app.services.distribuciones import distribucion_por_categoria_solucion, distribucion_por_ventana
from app.services.embeddings import asegurar_embeddings
from app.services.exportar_seguro import celda_segura
from app.services.excel_resumen import agregar_hoja_resumen
from app.services.gemini import GeminiError, extraer_mesa, gemini_configurado
from app.services.jefes import DANIO_POR_ACCION, DANIO_POR_LOGRO, danar_jefe
from app.services.logros import evaluar_logros
from app.services.reporte_semanal_export import (
    filas_reporte,
    generar_pdf_reporte,
    generar_pptx_reporte,
    generar_xlsx_reporte,
    titulo_reporte,
)
from app.services.auth import get_usuario_actual
from app.services.rpg import XP_POR_ACCION, XP_POR_LOGRO
from app.services.xp import otorgar_xp
from app.services.semanas import rango_semana, semana_de

router = APIRouter(prefix="/mesas", tags=["mesas"], dependencies=[Depends(get_usuario_actual)])

# Mesa.ventana/categoria/solicitante/resolutor son lazy="joined" en el modelo
# (siempre se cargan con JOIN): no hace falta pedir estas relaciones aquí. Un
# selectinload explícito anularía ese default y forzaría un round-trip extra
# por relación en cada consulta, más lento con miles de filas.


async def _obtener_mesa(session: AsyncSession, mesa_id: int) -> Mesa:
    stmt = select(Mesa).where(Mesa.id == mesa_id)
    return (await session.execute(stmt)).scalar_one()


def _ya_cerrada(mesa) -> bool:
    return mesa.fecha_cierre_real is not None


def _resolver_catalogo(catalogo: list, id_elegido: int | None):
    if not id_elegido:
        return None
    return next((c for c in catalogo if c.id == id_elegido), None)


async def _otorgar_xp_cierre(session: AsyncSession, mesa: Mesa, logros: list[str]) -> None:
    """Mismo XP sin importar si la mesa se cerró vía /cerrar o vía /editar."""
    cantidad = XP_POR_ACCION + XP_POR_LOGRO * len(logros)
    await otorgar_xp(
        session, mesa.resolutor.nombre, cantidad, "mesa_cerrada",
        usuario_id_directo=mesa.resolutor.usuario_id,
    )
    danio = DANIO_POR_ACCION + DANIO_POR_LOGRO * len(logros)
    await danar_jefe(session, semana_de(date.today()), danio)


@router.post("", response_model=MesaOut, status_code=201)
async def crear_mesa(payload: MesaCreate, session: AsyncSession = Depends(get_session)):
    mesa = Mesa(**payload.model_dump())
    session.add(mesa)
    try:
        await session.commit()
    except IntegrityError:
        await session.rollback()
        raise HTTPException(409, f"Código '{payload.codigo}' ya existe")
    mesa = await _obtener_mesa(session, mesa.id)
    logros = await evaluar_logros(session, mesa)
    await otorgar_xp(
        session, mesa.resolutor.nombre, XP_POR_ACCION, "mesa_creada",
        usuario_id_directo=mesa.resolutor.usuario_id,
    )
    await danar_jefe(session, semana_de(date.today()), DANIO_POR_ACCION)
    return MesaOut.model_validate(mesa).model_copy(update={"logros": logros})


@router.post("/extraer-imagen", response_model=ExtraccionMesa)
async def extraer_imagen_mesa(imagen: UploadFile = File(...), session: AsyncSession = Depends(get_session)):
    if not gemini_configurado():
        raise HTTPException(400, "Gemini no está configurado (falta GEMINI_API_KEY en el backend)")
    if not imagen.content_type or not imagen.content_type.startswith("image/"):
        raise HTTPException(400, "El archivo debe ser una imagen")

    contenido = await imagen.read()
    if len(contenido) > 10_000_000:
        raise HTTPException(400, "Imagen demasiado grande (máx 10MB)")

    solicitantes = (await session.execute(select(SolicitanteMesa).order_by(SolicitanteMesa.nombre))).scalars().all()

    try:
        extraido = await extraer_mesa(contenido, {"solicitantes": solicitantes})
    except GeminiError as exc:
        raise HTTPException(502, str(exc)) from exc

    fecha_carga = None
    if extraido.get("fecha_carga"):
        try:
            fecha_carga = datetime.fromisoformat(extraido["fecha_carga"])
        except ValueError:
            fecha_carga = None

    return ExtraccionMesa(
        codigo=extraido.get("codigo"),
        titulo=extraido.get("titulo"),
        fecha_carga=fecha_carga,
        descripcion=extraido.get("descripcion"),
        solicitante=_resolver_catalogo(solicitantes, extraido.get("solicitante_id")),
    )


@router.post("/{mesa_id}/cerrar", response_model=MesaOut)
async def cerrar_mesa(mesa_id: int, payload: MesaCerrar, session: AsyncSession = Depends(get_session)):
    mesa = await session.get(Mesa, mesa_id)
    if mesa is None:
        raise HTTPException(404, "Mesa no encontrada")
    if _ya_cerrada(mesa):
        raise HTTPException(400, "La mesa ya está cerrada")

    mesa.ventana_id = payload.ventana_id
    mesa.solucion = payload.solucion
    mesa.tipo_solucion = payload.tipo_solucion
    mesa.fecha_cierre_real = payload.fecha_cierre_real
    await session.commit()
    mesa = await _obtener_mesa(session, mesa_id)
    logros = await evaluar_logros(session, mesa)
    await _otorgar_xp_cierre(session, mesa, logros)
    return MesaOut.model_validate(mesa).model_copy(update={"logros": logros})


def _campos_a_actualizar(payload: MesaUpdate) -> dict:
    return payload.model_dump(exclude_unset=True)


@router.post("/{mesa_id}/editar", response_model=MesaOut)
async def editar_mesa(mesa_id: int, payload: MesaUpdate, session: AsyncSession = Depends(get_session)):
    mesa = await session.get(Mesa, mesa_id)
    if mesa is None:
        raise HTTPException(404, "Mesa no encontrada")

    recien_cerrada = mesa.fecha_cierre_real is None and payload.fecha_cierre_real is not None

    for campo, valor in _campos_a_actualizar(payload).items():
        setattr(mesa, campo, valor)

    try:
        await session.commit()
    except IntegrityError:
        await session.rollback()
        raise HTTPException(409, f"Código '{payload.codigo}' ya existe")

    mesa = await _obtener_mesa(session, mesa_id)
    logros = await evaluar_logros(session, mesa) if recien_cerrada else []
    if recien_cerrada:
        await _otorgar_xp_cierre(session, mesa, logros)
    return MesaOut.model_validate(mesa).model_copy(update={"logros": logros})


@router.delete("/{mesa_id}", status_code=204)
async def eliminar_mesa(mesa_id: int, session: AsyncSession = Depends(get_session)):
    mesa = await session.get(Mesa, mesa_id)
    if mesa is None:
        raise HTTPException(404, "Mesa no encontrada")
    await session.delete(mesa)
    await session.commit()


def _aplicar_filtros(
    stmt, *, categoria_id, solicitante_id, resolutor_id, ventana_id, semana, fecha_desde, fecha_hasta,
    buscar, estado, prioridad=None, destacada=None, buscar_solucion=None, con_solucion=None,
    tipo_solucion=None,
):
    if prioridad:
        stmt = stmt.where(Mesa.prioridad.is_(True))
    if destacada:
        stmt = stmt.where(Mesa.destacada.is_(True))
    if categoria_id:
        stmt = stmt.where(Mesa.categoria_id == categoria_id)
    if solicitante_id:
        stmt = stmt.where(Mesa.solicitante_id == solicitante_id)
    if resolutor_id:
        stmt = stmt.where(Mesa.resolutor_id == resolutor_id)
    if ventana_id:
        stmt = stmt.where(Mesa.ventana_id == ventana_id)
    if semana:
        stmt = stmt.where(Mesa.semana == semana)
    if fecha_desde:
        stmt = stmt.where(Mesa.fecha_carga >= fecha_desde)
    if fecha_hasta:
        # fecha_carga ahora incluye hora: <= fecha_hasta excluiría cualquier mesa
        # cargada después de medianoche de ese mismo día. Se compara contra el
        # inicio del día siguiente para incluir el día completo.
        stmt = stmt.where(Mesa.fecha_carga < fecha_hasta + timedelta(days=1))
    if estado == "abierta":
        stmt = stmt.where(Mesa.fecha_cierre_real.is_(None))
    elif estado == "cerrada":
        stmt = stmt.where(Mesa.fecha_cierre_real.isnot(None))
    if buscar:
        like = f"%{buscar}%"
        stmt = stmt.where(
            Mesa.descripcion.ilike(like)
            | Mesa.titulo.ilike(like)
            | Mesa.codigo.ilike(like)
            | Mesa.solucion.ilike(like)
        )
    if buscar_solucion:
        stmt = stmt.where(Mesa.solucion.ilike(f"%{buscar_solucion}%"))
    if con_solucion:
        stmt = stmt.where(Mesa.solucion.isnot(None)).where(Mesa.solucion != "")
    if tipo_solucion:
        stmt = stmt.where(Mesa.tipo_solucion == tipo_solucion)
    return stmt


@router.get("", response_model=PaginaMesas)
async def listar_mesas(
    session: AsyncSession = Depends(get_session),
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=500),
    categoria_id: int | None = None,
    solicitante_id: int | None = None,
    resolutor_id: int | None = None,
    ventana_id: int | None = None,
    semana: str | None = None,
    fecha_desde: date | None = None,
    fecha_hasta: date | None = None,
    buscar: str | None = None,
    estado: str | None = None,
    prioridad: bool | None = None,
    destacada: bool | None = None,
    buscar_solucion: str | None = None,
    con_solucion: bool | None = None,
    tipo_solucion: str | None = None,
):
    base = _aplicar_filtros(
        select(Mesa),
        categoria_id=categoria_id, solicitante_id=solicitante_id, resolutor_id=resolutor_id,
        ventana_id=ventana_id, semana=semana, fecha_desde=fecha_desde, fecha_hasta=fecha_hasta,
        buscar=buscar, estado=estado, prioridad=prioridad, destacada=destacada,
        buscar_solucion=buscar_solucion, con_solucion=con_solucion, tipo_solucion=tipo_solucion,
    )
    total = (await session.execute(select(func.count()).select_from(base.subquery()))).scalar_one()

    stmt = (
        base.order_by(
            # abiertas primero; entre las cerradas, agrupadas por día de cierre
            # (más reciente primero); dentro del día, por fecha de creación.
            case((Mesa.fecha_cierre_real.is_(None), 0), else_=1),
            Mesa.fecha_cierre_real.desc(),
            Mesa.fecha_carga.desc(),
            Mesa.id.desc(),
        )
        .offset((page - 1) * page_size)
        .limit(page_size)
    )
    items = (await session.execute(stmt)).scalars().all()
    return PaginaMesas(total=total, items=[MesaOut.model_validate(i) for i in items])


async def _mesas_de_semana(session: AsyncSession, semana: str) -> list[Mesa]:
    stmt = (
        select(Mesa)
        .where(Mesa.semana == semana)
        .order_by(Mesa.fecha_carga.desc(), Mesa.id.desc())
    )
    return (await session.execute(stmt)).scalars().all()


async def _mesas_cerradas_de_semana(session: AsyncSession, semana: str) -> list[Mesa]:
    """A diferencia de _mesas_de_semana (que agrupa por semana de CREACIÓN vía la
    columna generada Mesa.semana), el resumen y los reportes agrupan por semana
    de CIERRE: una mesa levantada un viernes y cerrada el lunes siguiente
    pertenece a la semana del lunes, no a la de creación."""
    lunes, _ = rango_semana(semana)
    inicio = datetime.combine(lunes, time.min)
    fin = inicio + timedelta(days=7)
    stmt = (
        select(Mesa)
        .where(Mesa.fecha_cierre_real >= inicio, Mesa.fecha_cierre_real < fin)
        .order_by(Mesa.fecha_cierre_real.desc(), Mesa.id.desc())
    )
    return (await session.execute(stmt)).scalars().all()


@router.get("/panel", response_model=PanelMesasKPIs)
async def panel_mesas(fecha: date | None = Query(None), session: AsyncSession = Depends(get_session)):
    semana = semana_de(fecha or date.today())
    mesas_semana = await _mesas_cerradas_de_semana(session, semana)

    por_dia: dict[str, int] = {}
    for m in mesas_semana:
        dia = m.fecha_cierre_real.date().isoformat()
        por_dia[dia] = por_dia.get(dia, 0) + 1

    volumen_diario = [{"fecha": k, "total": v} for k, v in sorted(por_dia.items())]
    distribucion_ventana = [{"ventana": k, "total": v} for k, v in distribucion_por_ventana(mesas_semana)]
    distribucion_categoria_solucion = [
        {"categoria_solucion": k, "total": v} for k, v in distribucion_por_categoria_solucion(mesas_semana)
    ]

    return PanelMesasKPIs(
        semana=semana,
        total_semana=len(mesas_semana),
        volumen_diario=volumen_diario,
        distribucion_ventana=distribucion_ventana,
        distribucion_categoria_solucion=distribucion_categoria_solucion,
        recientes=[MesaOut.model_validate(m) for m in mesas_semana[:10]],
    )


@router.get("/reporte", response_model=ReporteMesasSemanal)
async def reporte_mesas_semanal(semana: str = Query(...), session: AsyncSession = Depends(get_session)):
    mesas_semana = await _mesas_cerradas_de_semana(session, semana)
    if not mesas_semana:
        return ReporteMesasSemanal(semana=semana, total=0, por_categoria={}, por_solicitante={}, mesas=[])

    por_categoria: dict[str, int] = {}
    por_solicitante: dict[str, int] = {}
    for m in mesas_semana:
        por_categoria[m.categoria.nombre] = por_categoria.get(m.categoria.nombre, 0) + 1
        por_solicitante[m.solicitante.nombre] = por_solicitante.get(m.solicitante.nombre, 0) + 1

    return ReporteMesasSemanal(
        semana=semana, total=len(mesas_semana),
        por_categoria=por_categoria, por_solicitante=por_solicitante,
        mesas=[MesaOut.model_validate(m) for m in mesas_semana],
    )


@router.get("/reporte/exportar")
async def exportar_reporte_semanal(
    semana: str = Query(...),
    formato: str = Query("xlsx", pattern="^(xlsx|pptx|pdf)$"),
    session: AsyncSession = Depends(get_session),
):
    cerradas = await _mesas_cerradas_de_semana(session, semana)

    titulo = titulo_reporte(semana)
    filas = filas_reporte(cerradas)
    graficas = {
        "Por categoría de solución": distribucion_por_categoria_solucion(cerradas),
        "Por ventana": distribucion_por_ventana(cerradas),
    }
    nombre_base = f"reporte_semanal_{semana.replace(' ', '').replace('-', '_')}"

    if formato == "xlsx":
        out = generar_xlsx_reporte(titulo, filas, graficas)
        media = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    elif formato == "pptx":
        out = generar_pptx_reporte(titulo, filas, graficas)
        media = "application/vnd.openxmlformats-officedocument.presentationml.presentation"
    else:
        out = generar_pdf_reporte(titulo, filas, graficas)
        media = "application/pdf"

    return StreamingResponse(
        out, media_type=media, headers={"Content-Disposition": f"attachment; filename={nombre_base}.{formato}"}
    )


async def _descripciones_de_semana(session: AsyncSession, semana: str) -> list[Row]:
    """Solo id/descripcion/embedding: temas-frecuentes no necesita categoría/solicitante/resolutor/ventana."""
    stmt = select(Mesa.id, Mesa.descripcion, Mesa.embedding).where(Mesa.semana == semana)
    return list((await session.execute(stmt)).all())


@router.get("/temas-frecuentes", response_model=list[GrupoSoporte])
async def temas_frecuentes_mesas(
    semana: str = Query(...),
    umbral: float = Query(0.86, ge=0.5, le=0.99),
    top: int = Query(10, ge=1, le=50),
    session: AsyncSession = Depends(get_session),
):
    if not gemini_configurado():
        raise HTTPException(400, "Gemini no está configurado (falta GEMINI_API_KEY en el backend)")

    filas = await _descripciones_de_semana(session, semana)
    if not filas:
        return []

    try:
        embeddings = await asegurar_embeddings(session, filas, Mesa)
    except GeminiError as exc:
        raise HTTPException(502, str(exc)) from exc

    grupos = [
        GrupoSoporte(tema=tema_representativo([filas[i].descripcion for i in idx]), cantidad=len(idx))
        for idx in agrupar_por_similitud(embeddings, umbral)
        if len(idx) >= 2
    ]
    return grupos[:top]


def _generar_xlsx_mesas(encabezados: list[str], mesas_filtradas: list[Mesa]) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Mesas"

    color_encabezado = "0F1B2D"
    color_fila_a = "16273D"
    color_fila_b = "1C3149"
    color_borde = "2A3F58"
    color_texto = "E8EDF2"

    fuente_encabezado = Font(name="Calibri", bold=True, color=color_texto, size=11)
    relleno_encabezado = PatternFill("solid", fgColor=color_encabezado)
    alineacion_encabezado = Alignment(horizontal="left", vertical="center")
    fuente_cuerpo = Font(name="Calibri", color=color_texto, size=11)
    borde_fino = Border(bottom=Side(style="thin", color=color_borde))

    ws.sheet_view.showGridLines = False

    ws.append(encabezados)
    for celda in ws[1]:
        celda.font = fuente_encabezado
        celda.fill = relleno_encabezado
        celda.alignment = alineacion_encabezado
    ws.row_dimensions[1].height = 20

    for i, m in enumerate(mesas_filtradas, start=2):
        ws.cell(i, 1, celda_segura(m.codigo))
        ws.cell(i, 2, celda_segura(m.titulo))
        ws.cell(i, 3, m.fecha_carga).number_format = "yyyy-mm-dd hh:mm"
        ws.cell(i, 4, celda_segura(m.categoria.nombre))
        ws.cell(i, 5, celda_segura(m.solicitante.nombre))
        ws.cell(i, 6, celda_segura(m.resolutor.nombre))
        ws.cell(i, 7, celda_segura(m.ventana.nombre if m.ventana else ""))
        ws.cell(i, 8, celda_segura(m.descripcion))
        ws.cell(i, 9, "Cerrada" if m.fecha_cierre_real else "Abierta")
        ws.cell(i, 10, celda_segura(m.solucion or "—"))
        relleno_fila = PatternFill("solid", fgColor=color_fila_b if i % 2 == 0 else color_fila_a)
        for col in range(1, len(encabezados) + 1):
            celda = ws.cell(i, col)
            celda.border = borde_fino
            celda.font = fuente_cuerpo
            celda.fill = relleno_fila
            celda.alignment = Alignment(vertical="center", wrap_text=col == 8)

    anchos = [12, 30, 14, 16, 18, 14, 14, 50, 10, 40]
    for col, ancho in enumerate(anchos, start=1):
        ws.column_dimensions[get_column_letter(col)].width = ancho

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(encabezados))}{max(len(mesas_filtradas) + 1, 1)}"

    filas_resumen = [
        SimpleNamespace(fecha=m.fecha_carga.date(), descripcion=m.descripcion, embedding=m.embedding)
        for m in mesas_filtradas
    ]
    agregar_hoja_resumen(wb, filas_resumen)

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out


@router.get("/export")
async def exportar_mesas(
    formato: str = Query("csv", pattern="^(csv|xlsx)$"),
    session: AsyncSession = Depends(get_session),
    categoria_id: int | None = None,
    solicitante_id: int | None = None,
    resolutor_id: int | None = None,
    ventana_id: int | None = None,
    semana: str | None = None,
    fecha_desde: date | None = None,
    fecha_hasta: date | None = None,
    buscar: str | None = None,
    estado: str | None = None,
    prioridad: bool | None = None,
    destacada: bool | None = None,
):
    stmt = _aplicar_filtros(
        select(Mesa),
        categoria_id=categoria_id, solicitante_id=solicitante_id, resolutor_id=resolutor_id,
        ventana_id=ventana_id, semana=semana, fecha_desde=fecha_desde, fecha_hasta=fecha_hasta,
        buscar=buscar, estado=estado, prioridad=prioridad, destacada=destacada,
    ).order_by(Mesa.fecha_carga.desc())

    mesas_filtradas = (await session.execute(stmt)).scalars().all()
    encabezados = ["Código", "Título", "Fecha carga", "Categoría", "Solicitante", "Resolutor", "Ventana", "Descripción", "Estado", "Solución"]

    if formato == "csv":
        filas = [
            [
                celda_segura(m.codigo), celda_segura(m.titulo), m.fecha_carga.strftime("%Y-%m-%d %H:%M"),
                celda_segura(m.categoria.nombre), celda_segura(m.solicitante.nombre),
                celda_segura(m.resolutor.nombre), celda_segura(m.ventana.nombre if m.ventana else ""),
                celda_segura(m.descripcion), "Cerrada" if m.fecha_cierre_real else "Abierta",
                celda_segura(m.solucion or ""),
            ]
            for m in mesas_filtradas
        ]
        buf = io.StringIO()
        writer = csv.writer(buf)
        writer.writerow(encabezados)
        writer.writerows(filas)
        return StreamingResponse(
            iter([buf.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=mesas.csv"},
        )

    out = _generar_xlsx_mesas(encabezados, mesas_filtradas)
    return StreamingResponse(
        out,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=mesas.xlsx"},
    )


@router.get("/{mesa_id}", response_model=MesaOut)
async def obtener_mesa(mesa_id: int, session: AsyncSession = Depends(get_session)):
    """Registrada al final: debe ir después de todas las rutas GET de path literal
    (/panel, /reporte, /temas-frecuentes, /export) para que Starlette no la use
    como match genérico antes de llegar a esas — mismo bug que causó un 405 en
    producción cuando /soportes-frecuentes no existía en el backend viejo."""
    try:
        mesa = await _obtener_mesa(session, mesa_id)
    except NoResultFound:
        raise HTTPException(404, "Mesa no encontrada") from None
    return MesaOut.model_validate(mesa)
