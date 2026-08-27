import csv
import io
from datetime import date

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import Row, func, select
from sqlalchemy.exc import IntegrityError
from sqlalchemy.ext.asyncio import AsyncSession

from app.database import get_session
from app.models import Agente, Empresa, Medio, Modulo, Registro, Sistema
from app.schemas import (
    ExtraccionRegistro,
    GrupoSoporte,
    PanelKPIs,
    PaginaRegistros,
    RegistroCreadoOut,
    RegistroCreate,
    RegistroOut,
    RegistroUpdate,
    ReporteSemanal,
)
from app.services.clustering import agrupar_por_similitud, tema_representativo
from app.services.embeddings import asegurar_embeddings
from app.services.excel_resumen import agregar_hoja_resumen
from app.services.exportar_seguro import celda_segura
from app.services.gemini import GeminiError, extraer_registro, gemini_configurado
from app.services.auth import get_usuario_actual
from app.services.bonus import aplicar_bono_y_critico, es_critico_soporte, porcentaje_bono_del_momento
from app.services.jefes import DANIO_POR_ACCION, danar_jefe
from app.services.logros import a_hora_local, evaluar_logros_registro
from app.services.rpg import XP_POR_ACCION
from app.services.tienda import CREDITOS_POR_ACCION, otorgar_creditos
from app.services.semanas import semana_de
from app.services.trello import TrelloError, crear_tarjeta, trello_configurado
from app.services.xp import otorgar_xp

router = APIRouter(prefix="/registros", tags=["registros"], dependencies=[Depends(get_usuario_actual)])


# Registro.empresa/sistema/medio/modulo/atendio son lazy="joined" en el
# modelo (siempre se cargan con JOIN): no hace falta pedir estas relaciones
# aquí. Un selectinload explícito anularía ese default y forzaría un
# round-trip extra por relación en cada consulta.


async def _get_registro(session: AsyncSession, registro_id: int) -> Registro:
    stmt = select(Registro).where(Registro.id == registro_id)
    result = await session.execute(stmt)
    return result.scalar_one()


def _campos_a_actualizar(payload: RegistroUpdate) -> dict:
    return payload.model_dump(exclude_unset=True)


@router.post("", response_model=RegistroCreadoOut, status_code=201)
async def crear_registro(payload: RegistroCreate, session: AsyncSession = Depends(get_session)):
    registro = Registro(
        fecha=payload.fecha,
        empresa_id=payload.empresa_id,
        sistema_id=payload.sistema_id,
        medio_id=payload.medio_id,
        modulo_id=payload.modulo_id,
        atendio_id=payload.atendio_id,
        descripcion=payload.descripcion,
    )
    session.add(registro)
    await session.commit()
    registro = await _get_registro(session, registro.id)

    momento_local = a_hora_local(registro.created_at)
    porcentaje_bono, _ = await porcentaje_bono_del_momento(session, momento_local)
    critico = await es_critico_soporte(session, momento_local)

    xp = aplicar_bono_y_critico(XP_POR_ACCION, porcentaje_bono, critico)
    await otorgar_xp(
        session, registro.atendio.nombre, xp, "registro_creado",
        usuario_id_directo=registro.atendio.usuario_id,
    )
    danio = aplicar_bono_y_critico(DANIO_POR_ACCION, porcentaje_bono, critico)
    await danar_jefe(session, semana_de(date.today()), danio, registro.atendio.nombre, "registro_creado")
    await otorgar_creditos(
        session, registro.atendio.nombre, CREDITOS_POR_ACCION, "registro_creado", semana_de(date.today()),
        usuario_id_directo=registro.atendio.usuario_id,
    )
    logros = await evaluar_logros_registro(session, registro)

    trello_ok = False
    trello_error = None
    if trello_configurado():
        try:
            titulo = f"{registro.empresa.nombre} · {registro.sistema.nombre} · {registro.modulo.nombre}"
            card_id = await crear_tarjeta(titulo, registro.descripcion)
            registro.trello_card_id = card_id
            await session.commit()
            registro = await _get_registro(session, registro.id)
            trello_ok = True
        except TrelloError as exc:
            trello_error = str(exc)

    return RegistroCreadoOut(
        registro=RegistroOut.model_validate(registro), trello_ok=trello_ok, trello_error=trello_error, logros=logros,
    )


@router.post("/{registro_id}/editar", response_model=RegistroOut)
async def editar_registro(registro_id: int, payload: RegistroUpdate, session: AsyncSession = Depends(get_session)):
    registro = await session.get(Registro, registro_id)
    if registro is None:
        raise HTTPException(404, "Registro no encontrado")

    for campo, valor in _campos_a_actualizar(payload).items():
        setattr(registro, campo, valor)

    try:
        await session.commit()
    except IntegrityError:
        await session.rollback()
        raise HTTPException(400, "Datos inválidos (revisa los catálogos seleccionados)")

    return await _get_registro(session, registro_id)


@router.post("/extraer-imagen", response_model=ExtraccionRegistro)
async def extraer_imagen(imagen: UploadFile = File(...), session: AsyncSession = Depends(get_session)):
    if not gemini_configurado():
        raise HTTPException(400, "Gemini no está configurado (falta GEMINI_API_KEY en el backend)")
    if not imagen.content_type or not imagen.content_type.startswith("image/"):
        raise HTTPException(400, "El archivo debe ser una imagen")

    contenido = await imagen.read()
    if len(contenido) > 10_000_000:
        raise HTTPException(400, "Imagen demasiado grande (máx 10MB)")

    catalogos = {}
    for clave, modelo in (
        ("empresas", Empresa), ("sistemas", Sistema), ("medios", Medio),
        ("modulos", Modulo), ("agentes", Agente),
    ):
        catalogos[clave] = (await session.execute(select(modelo).order_by(modelo.nombre))).scalars().all()

    try:
        extraido = await extraer_registro(contenido, catalogos)
    except GeminiError as exc:
        raise HTTPException(502, str(exc)) from exc

    def _match(clave: str, campo: str):
        id_elegido = extraido.get(campo)
        if not id_elegido:
            return None
        return next((c for c in catalogos[clave] if c.id == id_elegido), None)

    fecha = None
    if extraido.get("fecha"):
        try:
            fecha = date.fromisoformat(extraido["fecha"])
        except ValueError:
            fecha = None

    return ExtraccionRegistro(
        fecha=fecha,
        descripcion=extraido.get("descripcion"),
        empresa=_match("empresas", "empresa_id"),
        sistema=_match("sistemas", "sistema_id"),
        medio=_match("medios", "medio_id"),
        modulo=_match("modulos", "modulo_id"),
        atendio=_match("agentes", "atendio_id"),
    )


@router.post("/{registro_id}/reintentar-trello", response_model=RegistroCreadoOut)
async def reintentar_trello(registro_id: int, session: AsyncSession = Depends(get_session)):
    registro = await _get_registro(session, registro_id)
    trello_ok = False
    trello_error = None
    try:
        titulo = f"{registro.empresa.nombre} · {registro.sistema.nombre} · {registro.modulo.nombre}"
        card_id = await crear_tarjeta(titulo, registro.descripcion)
        registro.trello_card_id = card_id
        await session.commit()
        registro = await _get_registro(session, registro.id)
        trello_ok = True
    except TrelloError as exc:
        trello_error = str(exc)
    return RegistroCreadoOut(registro=RegistroOut.model_validate(registro), trello_ok=trello_ok, trello_error=trello_error)


@router.delete("/{registro_id}", status_code=204)
async def eliminar_registro(registro_id: int, session: AsyncSession = Depends(get_session)):
    registro = await session.get(Registro, registro_id)
    if registro is None:
        raise HTTPException(404, "Registro no encontrado")
    await session.delete(registro)
    await session.commit()


def _aplicar_filtros(stmt, *, empresa_id, sistema_id, medio_id, modulo_id, atendio_id, semana, fecha_desde, fecha_hasta, buscar):
    if empresa_id:
        stmt = stmt.where(Registro.empresa_id == empresa_id)
    if sistema_id:
        stmt = stmt.where(Registro.sistema_id == sistema_id)
    if medio_id:
        stmt = stmt.where(Registro.medio_id == medio_id)
    if modulo_id:
        stmt = stmt.where(Registro.modulo_id == modulo_id)
    if atendio_id:
        stmt = stmt.where(Registro.atendio_id == atendio_id)
    if semana:
        stmt = stmt.where(Registro.semana == semana)
    if fecha_desde:
        stmt = stmt.where(Registro.fecha >= fecha_desde)
    if fecha_hasta:
        stmt = stmt.where(Registro.fecha <= fecha_hasta)
    if buscar:
        like = f"%{buscar}%"
        stmt = (
            stmt.join(Empresa)
            .join(Sistema)
            .join(Medio)
            .join(Modulo)
            .join(Agente)
            .where(
                Registro.descripcion.ilike(like)
                | Empresa.nombre.ilike(like)
                | Sistema.nombre.ilike(like)
                | Medio.nombre.ilike(like)
                | Modulo.nombre.ilike(like)
                | Agente.nombre.ilike(like)
            )
        )
    return stmt


@router.get("", response_model=PaginaRegistros)
async def listar_registros(
    session: AsyncSession = Depends(get_session),
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=500),
    empresa_id: int | None = None,
    sistema_id: int | None = None,
    medio_id: int | None = None,
    modulo_id: int | None = None,
    atendio_id: int | None = None,
    semana: str | None = None,
    fecha_desde: date | None = None,
    fecha_hasta: date | None = None,
    buscar: str | None = None,
):
    base = _aplicar_filtros(
        select(Registro),
        empresa_id=empresa_id, sistema_id=sistema_id, medio_id=medio_id,
        modulo_id=modulo_id, atendio_id=atendio_id, semana=semana,
        fecha_desde=fecha_desde, fecha_hasta=fecha_hasta, buscar=buscar,
    )
    total = (await session.execute(select(func.count()).select_from(base.subquery()))).scalar_one()

    stmt = (
        base.order_by(Registro.fecha.desc(), Registro.id.desc())
        .offset((page - 1) * page_size)
        .limit(page_size)
    )
    items = (await session.execute(stmt)).scalars().all()
    return PaginaRegistros(total=total, items=[RegistroOut.model_validate(i) for i in items])


async def _registros_de_semana(session: AsyncSession, semana: str) -> list[Registro]:
    stmt = (
        select(Registro)
        .where(Registro.semana == semana)
        .order_by(Registro.fecha.desc(), Registro.id.desc())
    )
    return (await session.execute(stmt)).scalars().all()


async def _kpis_registros_de_semana(session: AsyncSession, semana: str):
    """Agrega en SQL en vez de traer toda la semana y sumar en Python — el
    panel se visita seguido y ahora también se refresca solo cada 30s."""
    total = await session.scalar(select(func.count()).select_from(Registro).where(Registro.semana == semana))

    por_sistema = (
        await session.execute(
            select(Sistema.nombre, func.count())
            .select_from(Registro)
            .join(Sistema, Registro.sistema_id == Sistema.id)
            .where(Registro.semana == semana)
            .group_by(Sistema.nombre)
            .order_by(Sistema.nombre)
        )
    ).all()
    por_dia = (
        await session.execute(
            select(Registro.fecha, func.count())
            .where(Registro.semana == semana)
            .group_by(Registro.fecha)
            .order_by(Registro.fecha)
        )
    ).all()
    por_modulo = (
        await session.execute(
            select(Modulo.nombre, func.count())
            .select_from(Registro)
            .join(Modulo, Registro.modulo_id == Modulo.id)
            .where(Registro.semana == semana)
            .group_by(Modulo.nombre)
            .order_by(func.count().desc(), Modulo.nombre)
        )
    ).all()
    return total or 0, por_sistema, por_dia, por_modulo


async def _registros_recientes_de_semana(session: AsyncSession, semana: str, limite: int = 10) -> list[Registro]:
    stmt = (
        select(Registro)
        .where(Registro.semana == semana)
        .order_by(Registro.fecha.desc(), Registro.id.desc())
        .limit(limite)
    )
    return (await session.execute(stmt)).scalars().all()


@router.get("/panel", response_model=PanelKPIs)
async def panel(session: AsyncSession = Depends(get_session)):
    semana = semana_de(date.today())
    total, filas_sistema, filas_dia, filas_modulo = await _kpis_registros_de_semana(session, semana)
    recientes = await _registros_recientes_de_semana(session, semana)

    por_sistema = {nombre: cantidad for nombre, cantidad in filas_sistema}
    volumen_diario = [{"fecha": fecha.isoformat(), "total": cantidad} for fecha, cantidad in filas_dia]
    distribucion_modulo = [{"modulo": nombre, "total": cantidad} for nombre, cantidad in filas_modulo]

    return PanelKPIs(
        semana=semana,
        total_semana=total,
        por_sistema=por_sistema,
        volumen_diario=volumen_diario,
        distribucion_modulo=distribucion_modulo,
        recientes=[RegistroOut.model_validate(r) for r in recientes],
    )


@router.get("/reporte", response_model=ReporteSemanal)
async def reporte_semanal(semana: str = Query(...), session: AsyncSession = Depends(get_session)):
    registros = await _registros_de_semana(session, semana)
    if not registros:
        return ReporteSemanal(semana=semana, total=0, por_sistema={}, por_empresa={}, por_medio={}, registros=[])

    por_sistema: dict[str, int] = {}
    por_empresa: dict[str, int] = {}
    por_medio: dict[str, int] = {}
    for r in registros:
        por_sistema[r.sistema.nombre] = por_sistema.get(r.sistema.nombre, 0) + 1
        por_empresa[r.empresa.nombre] = por_empresa.get(r.empresa.nombre, 0) + 1
        por_medio[r.medio.nombre] = por_medio.get(r.medio.nombre, 0) + 1

    return ReporteSemanal(
        semana=semana, total=len(registros),
        por_sistema=por_sistema, por_empresa=por_empresa, por_medio=por_medio,
        registros=[RegistroOut.model_validate(r) for r in registros],
    )


async def _descripciones_de_semana(session: AsyncSession, semana: str) -> list[Row]:
    """Solo id/descripcion/embedding: este endpoint no necesita empresa/sistema/medio/módulo/agente."""
    stmt = select(Registro.id, Registro.descripcion, Registro.embedding).where(Registro.semana == semana)
    return list((await session.execute(stmt)).all())


@router.get("/soportes-frecuentes", response_model=list[GrupoSoporte])
async def soportes_frecuentes(
    semana: str = Query(...),
    umbral: float = Query(0.86, ge=0.5, le=0.99),
    top: int = Query(10, ge=1, le=50),
    session: AsyncSession = Depends(get_session),
):
    if not gemini_configurado():
        raise HTTPException(400, "Gemini no está configurado (falta GEMINI_API_KEY en el backend)")

    registros = await _descripciones_de_semana(session, semana)
    if not registros:
        return []

    try:
        embeddings = await asegurar_embeddings(session, registros, Registro)
    except GeminiError as exc:
        raise HTTPException(502, str(exc)) from exc

    grupos = [
        GrupoSoporte(tema=tema_representativo([registros[i].descripcion for i in idx]), cantidad=len(idx))
        for idx in agrupar_por_similitud(embeddings, umbral)
        if len(idx) >= 2
    ]
    return grupos[:top]


@router.get("/export")
async def exportar(
    formato: str = Query("csv", pattern="^(csv|xlsx)$"),
    session: AsyncSession = Depends(get_session),
    empresa_id: int | None = None,
    sistema_id: int | None = None,
    medio_id: int | None = None,
    modulo_id: int | None = None,
    atendio_id: int | None = None,
    semana: str | None = None,
    fecha_desde: date | None = None,
    fecha_hasta: date | None = None,
    buscar: str | None = None,
):
    stmt = _aplicar_filtros(
        select(Registro),
        empresa_id=empresa_id, sistema_id=sistema_id, medio_id=medio_id,
        modulo_id=modulo_id, atendio_id=atendio_id, semana=semana,
        fecha_desde=fecha_desde, fecha_hasta=fecha_hasta, buscar=buscar,
    ).order_by(Registro.fecha.desc())

    registros = (await session.execute(stmt)).scalars().all()
    encabezados = ["Fecha", "Semana", "Empresa", "Sistema", "Medio", "Módulo", "Atendió", "Descripción", "Tarjeta Trello"]

    if formato == "csv":
        filas = [
            [
                r.fecha.isoformat(), r.semana, celda_segura(r.empresa.nombre), celda_segura(r.sistema.nombre),
                celda_segura(r.medio.nombre), celda_segura(r.modulo.nombre), celda_segura(r.atendio.nombre),
                celda_segura(r.descripcion), celda_segura(r.trello_card_id or ""),
            ]
            for r in registros
        ]
        buf = io.StringIO()
        writer = csv.writer(buf)
        writer.writerow(encabezados)
        writer.writerows(filas)
        return StreamingResponse(
            iter([buf.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=bitacora.csv"},
        )

    out = _generar_xlsx(encabezados, registros)
    return StreamingResponse(
        out,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=bitacora.xlsx"},
    )


def _generar_xlsx(encabezados: list[str], registros: list[Registro]) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Bitácora"

    color_encabezado = "0F1B2D"
    color_fila_a = "16273D"
    color_fila_b = "1C3149"
    color_borde = "2A3F58"
    color_texto = "E8EDF2"

    fuente_encabezado = Font(name="Calibri", bold=True, color=color_texto, size=11)
    relleno_encabezado = PatternFill("solid", fgColor=color_encabezado)
    alineacion_encabezado = Alignment(horizontal="left", vertical="center")
    fuente_cuerpo = Font(name="Calibri", color=color_texto, size=11)
    borde_fino = Border(bottom=Side(style="thin", color=color_borde))

    ws.sheet_view.showGridLines = False

    ws.append(encabezados)
    for celda in ws[1]:
        celda.font = fuente_encabezado
        celda.fill = relleno_encabezado
        celda.alignment = alineacion_encabezado
    ws.row_dimensions[1].height = 20

    for i, r in enumerate(registros, start=2):
        ws.cell(i, 1, r.fecha).number_format = "yyyy-mm-dd"
        ws.cell(i, 2, r.semana)
        ws.cell(i, 3, celda_segura(r.empresa.nombre))
        ws.cell(i, 4, celda_segura(r.sistema.nombre))
        ws.cell(i, 5, celda_segura(r.medio.nombre))
        ws.cell(i, 6, celda_segura(r.modulo.nombre))
        ws.cell(i, 7, celda_segura(r.atendio.nombre))
        ws.cell(i, 8, celda_segura(r.descripcion))
        ws.cell(i, 9, celda_segura(r.trello_card_id or "—"))
        relleno_fila = PatternFill("solid", fgColor=color_fila_b if i % 2 == 0 else color_fila_a)
        for col in range(1, len(encabezados) + 1):
            celda = ws.cell(i, col)
            celda.border = borde_fino
            celda.font = fuente_cuerpo
            celda.fill = relleno_fila
            celda.alignment = Alignment(vertical="center", wrap_text=col == 8)

    anchos = [12, 15, 22, 14, 14, 22, 18, 50, 16]
    for col, ancho in enumerate(anchos, start=1):
        ws.column_dimensions[get_column_letter(col)].width = ancho

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(encabezados))}{max(len(registros) + 1, 1)}"

    agregar_hoja_resumen(wb, registros)

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out


@router.get("/{registro_id}", response_model=RegistroOut)
async def obtener_registro(registro_id: int, session: AsyncSession = Depends(get_session)):
    stmt = select(Registro).where(Registro.id == registro_id)
    registro = (await session.execute(stmt)).scalar_one_or_none()
    if registro is None:
        raise HTTPException(404, "Registro no encontrado")
    return registro
