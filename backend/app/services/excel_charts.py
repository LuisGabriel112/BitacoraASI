from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font
from openpyxl.worksheet.worksheet import Worksheet


def escribir_tabla_conteo(ws: Worksheet, titulo: str, fila_inicio: int, datos: list[tuple[str, int]]) -> tuple[int, int]:
    """Escribe título + encabezados + filas (etiqueta, cantidad). Devuelve (fila_datos_inicio, fila_siguiente_bloque)."""
    ws.cell(fila_inicio, 1, titulo).font = Font(bold=True)
    ws.cell(fila_inicio + 1, 1, "Etiqueta")
    ws.cell(fila_inicio + 1, 2, "Cantidad")
    fila_datos = fila_inicio + 2
    for i, (etiqueta, cantidad) in enumerate(datos, start=fila_datos):
        ws.cell(i, 1, etiqueta)
        ws.cell(i, 2, cantidad)
    return fila_datos, fila_datos + len(datos)


def grafica_barras_conteo(
    ws: Worksheet, titulo: str, fila_datos: int, num_filas: int, ancla: str, color_barra: str | None = None
) -> None:
    chart = BarChart()
    chart.title = titulo
    chart.y_axis.title = "Cantidad"
    datos = Reference(ws, min_col=2, min_row=fila_datos, max_row=fila_datos + num_filas - 1)
    categorias = Reference(ws, min_col=1, min_row=fila_datos, max_row=fila_datos + num_filas - 1)
    chart.add_data(datos, titles_from_data=False)
    chart.set_categories(categorias)
    if color_barra:
        chart.series[0].graphicalProperties.solidFill = color_barra
    ws.add_chart(chart, ancla)
