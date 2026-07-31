import io
import zipfile
from datetime import date
from types import SimpleNamespace

from openpyxl import Workbook

from app.services import excel_resumen


def _registro(fecha: date, descripcion: str, embedding: list[float] | None = None) -> SimpleNamespace:
    return SimpleNamespace(fecha=fecha, descripcion=descripcion, embedding=embedding)


def _nombres_de_grafica(wb: Workbook) -> list[str]:
    buffer = io.BytesIO()
    wb.save(buffer)
    with zipfile.ZipFile(buffer) as zf:
        return [n for n in zf.namelist() if n.startswith("xl/charts/chart")]


def test_conteo_por_dia_agrupa_y_ordena_por_fecha():
    registros = [
        _registro(date(2026, 7, 24), "a"),
        _registro(date(2026, 7, 25), "b"),
        _registro(date(2026, 7, 24), "c"),
    ]

    assert excel_resumen._conteo_por_dia(registros) == [("2026-07-24", 2), ("2026-07-25", 1)]


def test_conteo_por_dia_vacio():
    assert excel_resumen._conteo_por_dia([]) == []


def test_temas_frecuentes_vacio_si_gemini_no_configurado(monkeypatch):
    monkeypatch.setattr(excel_resumen, "gemini_configurado", lambda: False)
    registros = [_registro(date(2026, 7, 24), "a", [1.0, 0.0])]

    assert excel_resumen._temas_frecuentes(registros) == []


def test_temas_frecuentes_vacio_si_falta_algun_embedding(monkeypatch):
    monkeypatch.setattr(excel_resumen, "gemini_configurado", lambda: True)
    registros = [
        _registro(date(2026, 7, 24), "a", [1.0, 0.0]),
        _registro(date(2026, 7, 24), "b", None),
    ]

    assert excel_resumen._temas_frecuentes(registros) == []


def test_temas_frecuentes_agrupa_por_similitud(monkeypatch):
    monkeypatch.setattr(excel_resumen, "gemini_configurado", lambda: True)
    registros = [
        _registro(date(2026, 7, 24), "Cambio contraseña", [0.9, 0.1]),
        _registro(date(2026, 7, 24), "Cambiar contraseña", [0.91, 0.09]),
        _registro(date(2026, 7, 24), "Usuario bloqueado", [0.1, 0.9]),
    ]

    assert excel_resumen._temas_frecuentes(registros) == [("Cambio contraseña", 2)]


def test_agregar_hoja_resumen_sin_registros_no_falla():
    wb = Workbook()

    excel_resumen.agregar_hoja_resumen(wb, [])

    assert "Resumen" in wb.sheetnames
    assert _nombres_de_grafica(wb) == []


def test_agregar_hoja_resumen_incluye_barras_y_pastel_cuando_hay_datos(monkeypatch):
    monkeypatch.setattr(excel_resumen, "gemini_configurado", lambda: True)
    wb = Workbook()
    registros = [
        _registro(date(2026, 7, 24), "Cambio contraseña", [0.9, 0.1]),
        _registro(date(2026, 7, 24), "Cambiar contraseña", [0.91, 0.09]),
        _registro(date(2026, 7, 25), "Usuario bloqueado", [0.1, 0.9]),
    ]

    excel_resumen.agregar_hoja_resumen(wb, registros)

    assert "Resumen" in wb.sheetnames
    assert len(_nombres_de_grafica(wb)) == 2


def test_agregar_hoja_resumen_omite_pastel_sin_temas_frecuentes(monkeypatch):
    monkeypatch.setattr(excel_resumen, "gemini_configurado", lambda: False)
    wb = Workbook()
    registros = [_registro(date(2026, 7, 24), "Cambio contraseña", None)]

    excel_resumen.agregar_hoja_resumen(wb, registros)

    assert len(_nombres_de_grafica(wb)) == 1
