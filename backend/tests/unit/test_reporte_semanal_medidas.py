from datetime import datetime
from io import BytesIO
from types import SimpleNamespace

from openpyxl import load_workbook

from app.services.reporte_semanal_export import (
    ENCABEZADOS,
    OPCIONES_MEDIDAS,
    filas_reporte,
    generar_xlsx_reporte,
    texto_medidas,
)

_FILAS = [
    ("TCK-001", "03/08/2026 09:00", "Se aplicó el fix", OPCIONES_MEDIDAS[0]),
    ("TCK-002", "04/08/2026 10:00", "Se reinició el servicio", OPCIONES_MEDIDAS[0]),
    ("TCK-003", "05/08/2026 11:00", "Se corrigió el registro", OPCIONES_MEDIDAS[0]),
]


def _xlsx(filas=_FILAS):
    return load_workbook(BytesIO(generar_xlsx_reporte("Reporte de prueba", filas, {}).getvalue()))


def test_ultima_columna_se_llama_medidas_para_disminuir_el_impacto():
    assert ENCABEZADOS[-1] == "Medidas realizadas para disminuir el impacto"


def test_xlsx_columna_medidas_tiene_lista_desplegable():
    wb = _xlsx()
    validaciones = wb["Reporte"].data_validations.dataValidation

    assert len(validaciones) == 1
    assert validaciones[0].type == "list"


def test_opciones_de_la_lista_son_las_dos_exactas():
    assert OPCIONES_MEDIDAS == (
        "Ninguna.",
        "En progreso la validación de la correctez de la información ingresada por el proveedor. "
        "Esto determinará la acción correctiva: a) ajustar la validación para rechazar la info o  "
        "b) ajustar el sistema para aceptarla y procesarla.",
    )


def test_validacion_cubre_todas_las_filas_de_datos():
    wb = _xlsx()
    dv = wb["Reporte"].data_validations.dataValidation[0]

    assert str(dv.sqref) == "D4:D6"  # encabezado en fila 3, 3 filas de datos: 4,5,6


def test_sin_filas_no_agrega_validacion():
    wb = _xlsx(filas=[])
    validaciones = wb["Reporte"].data_validations.dataValidation

    assert len(validaciones) == 0


def test_texto_medidas_sin_marcar_es_ninguna():
    assert texto_medidas(False) == OPCIONES_MEDIDAS[0]


def test_texto_medidas_marcada_es_el_texto_largo():
    assert texto_medidas(True) == OPCIONES_MEDIDAS[1]


def _mesa(codigo: str, medidas_impacto: bool) -> SimpleNamespace:
    return SimpleNamespace(
        codigo=codigo,
        fecha_cierre_real=datetime(2026, 8, 4, 10, 0),
        solucion="listo",
        medidas_impacto=medidas_impacto,
    )


def test_filas_reporte_sin_checkbox_marcado_pone_ninguna():
    filas = filas_reporte([_mesa("TCK-010", medidas_impacto=False)])

    assert filas[0][3] == "Ninguna."


def test_filas_reporte_con_checkbox_marcado_pone_texto_largo():
    filas = filas_reporte([_mesa("TCK-011", medidas_impacto=True)])

    assert filas[0][3] == OPCIONES_MEDIDAS[1]


def test_xlsx_columna_medidas_trae_el_texto_de_cada_fila_no_solo_la_lista():
    filas = [
        ("TCK-001", "03/08/2026 09:00", "Se aplicó el fix", OPCIONES_MEDIDAS[0]),
        ("TCK-002", "04/08/2026 10:00", "Se reinició el servicio", OPCIONES_MEDIDAS[1]),
    ]
    wb = _xlsx(filas=filas)
    ws = wb["Reporte"]

    assert ws["D4"].value == OPCIONES_MEDIDAS[0]
    assert ws["D5"].value == OPCIONES_MEDIDAS[1]


def test_xlsx_columna_medidas_usa_fuente_tamano_9():
    wb = _xlsx()
    ws = wb["Reporte"]

    assert ws["D4"].font.size == 9


def test_xlsx_otras_columnas_mantienen_fuente_tamano_10():
    wb = _xlsx()
    ws = wb["Reporte"]

    assert ws["A4"].font.size == 10
    assert ws["C4"].font.size == 10
